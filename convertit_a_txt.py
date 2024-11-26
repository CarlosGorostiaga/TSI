import os
import re
from openpyxl import load_workbook
from datetime import datetime

# Define la ruta del directorio y obtén el archivo xls más reciente que comienza con "IDO"
directorio = "C:/CargaFotos"
archivos_xls = [f for f in os.listdir(directorio) if f.startswith("IDO") and f.endswith(".xlsx")]
ruta_xls = max([os.path.join(directorio, f) for f in archivos_xls], key=os.path.getmtime)

# Define la ruta del archivo de destino basado en el archivo de origen pero con extensión .txt
ruta_txt = os.path.join(directorio, os.path.splitext(os.path.basename(ruta_xls))[0] + ".txt")

# Cargar el archivo de Excel
workbook = load_workbook(ruta_xls, data_only=True)
worksheet = workbook.active

# Define las letras de las columnas que deseas extraer
columnas = ["A", "P", "C", "D", "E", "F", "L", "M", "T"]

# Recupera la última fila con datos
ultima_fila = worksheet.max_row

# Procesa las filas desde la segunda fila (para omitir la fila A1) y extrae los datos de las columnas seleccionadas
resultado = []
for i in range(2, ultima_fila + 1):
    fila = []
    for col in columnas:
        valor = worksheet[f"{col}{i}"].value

        # Verificar si la columna es "P" y está vacía
        if col == "P" and (valor is None or valor == ""):
            valor = "00000000" # Valor por defecto en caso de ausencia 

        # Transformar fechas y horas según las columnas
        if col in ["C", "D"] and valor:
            # Convertir la fecha de YYYY.MM.DD a DDMMYY
            try:
                valor = datetime.strptime(str(valor), "%Y.%m.%d").strftime("%d%m%y")
            except ValueError:
                valor = valor
        elif col in ["E", "F"] and valor:
            # Eliminar los segundos y los dos puntos de la hora
            valor = re.sub(r"(\d{2}):(\d{2}):\d{2}", r"\1\2", str(valor))
        elif col == "L" and valor:
            # Procesar latitud
            match = re.match(r"^(\d+)\.(\d{5})", str(valor))
            if match:
                entero = int(match.group(1))
                decimal = match.group(2)
                valor = f"{entero:03}{decimal}" if entero < 100 else f"{entero}{decimal}"
            else:
                valor = "00000000" # Valor por defecto en caso de ausencia 
        elif col == "M" and valor:
            # Procesar longitud
            match = re.match(r"^-?(\d+)\.(\d{5})", str(valor))
            if match:
                entero = int(match.group(1))
                decimal = match.group(2)
                valor = f"0-{entero}{decimal}" if valor.startswith("-") else f"0{entero}{decimal}"
            else:
                valor = "00000000" # Valor por defecto en caso de ausencia ...  

        fila.append(str(valor) if valor is not None else "")
    
    # Unir la fila procesada en una sola cadena
    resultado.append("".join(fila).replace("\n", "").replace(".", ""))

# Guarda el resultado en un archivo TXT
with open(ruta_txt, "w") as file:
    file.write("\n".join(resultado))

# Cierra el archivo de Excel
workbook.close()

# Mueve el archivo Excel a la carpeta C:\CargaFotosSeg
ruta_destino = os.path.join("C:/CargaFotosSeg", os.path.basename(ruta_xls))
os.rename(ruta_xls, ruta_destino) # Renombramos el archivo excel 




         

        



